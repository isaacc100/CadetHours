# excel_exporter.py
import openpyxl
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from PySide6.QtWidgets import QMessageBox

def export_to_excel(path, entries, parent_widget=None):
    """
    Export hourtracker entries to Excel with two sheets:
    - 'Hours' sheet: raw entries (date, type, hours, travel, total)
    - 'Summary' sheet: totals, monthly avg, year-to-date, charts
    """
    try:
        wb = openpyxl.Workbook()
        # --- Hours sheet ---
        ws = wb.active
        ws.title = "Hours"
        headers = ["Date", "Type", "Hours", "Travel", "Total"]
        ws.append(headers)

        for entry in entries:
            id_, date, type_, hours, travel = entry
            total = hours + travel
            ws.append([date, type_, hours, travel, total])
            
        # Set column widths
        column_widths = [15, 20, 10, 12, 10]  # Adjust as needed
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
            
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # --- Summary sheet ---
        summary_ws = wb.create_sheet(title="Summary")

        # Calculate monthly totals, averages and YTD
        monthly_hours = defaultdict(float)
        monthly_counts = defaultdict(int)
        type_hours = defaultdict(float)
        type_travel = defaultdict(float)

        for entry in entries:
            _, date_str, type_, hours, travel = entry
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            month_key = date_obj.strftime("%Y-%m")
            monthly_hours[month_key] += hours + travel
            monthly_counts[month_key] += 1
            type_hours[type_] += hours
            type_travel[type_] += travel

        # Write monthly summary table
        summary_ws.append(["Month", "Total Hours (incl. travel)", "Entries Count", "Average Hours per Entry"])
        for month in sorted(monthly_hours.keys()):
            total = monthly_hours[month]
            count = monthly_counts[month]
            avg = total / count if count else 0
            summary_ws.append([month, total, count, avg])

        # Calculate YTD (year-to-date) totals
        current_year = datetime.now().year
        ytd_total = 0
        for month in monthly_hours:
            year = int(month.split("-")[0])
            if year == current_year:
                ytd_total += monthly_hours[month]

        # Add YTD summary below monthly table
        ytd_row = summary_ws.max_row + 2
        summary_ws.cell(row=ytd_row, column=1, value=f"Year-to-date total hours for {current_year}")
        summary_ws.cell(row=ytd_row, column=2, value=ytd_total)
        summary_ws.cell(row=ytd_row, column=1).font = Font(bold=True)
        summary_ws.cell(row=ytd_row, column=2).font = Font(bold=True)

        # Create bar chart for monthly total hours
        bar_chart = BarChart()
        bar_chart.title = "Monthly Total Hours"
        bar_chart.y_axis.title = "Hours"
        bar_chart.x_axis.title = "Month"

        data = Reference(summary_ws, min_col=2, min_row=1, max_row=summary_ws.max_row-3)
        cats = Reference(summary_ws, min_col=1, min_row=2, max_row=summary_ws.max_row-3)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        summary_ws.add_chart(bar_chart, "F2")

        # Write type summary for pie chart
        type_start_row = summary_ws.max_row + 2
        summary_ws.cell(row=type_start_row, column=1, value="Type")
        summary_ws.cell(row=type_start_row, column=2, value="Hours")
        summary_ws.cell(row=type_start_row, column=3, value="Travel")
        for i, (t, h) in enumerate(type_hours.items(), start=type_start_row + 1):
            summary_ws.cell(row=i, column=1, value=t)
            summary_ws.cell(row=i, column=2, value=h)
            summary_ws.cell(row=i, column=3, value=type_travel[t])

        # Pie chart for type hours (excluding travel)
        pie = PieChart()
        pie.title = "Hours by Type (excluding travel)"
        labels = Reference(summary_ws, min_col=1, min_row=type_start_row + 1, max_row=type_start_row + len(type_hours))
        data = Reference(summary_ws, min_col=2, min_row=type_start_row, max_row=type_start_row + len(type_hours))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        summary_ws.add_chart(pie, "F20")

        wb.save(path)
        if parent_widget:
            QMessageBox.information(parent_widget, "Export Complete", f"Excel file saved to:\n{path}")

    except Exception as e:
        if parent_widget:
            QMessageBox.critical(parent_widget, "Error", f"Could not save Excel file:\n{e}")
        else:
            print(f"Error exporting Excel: {e}")
